#!/usr/bin/env python3
"""
Extract all sheets from the PeRef Data Dictionary xlsx to CSV files.
Output goes to input/data-dictionary/ with one CSV per sheet.

If the plaintext xlsx is absent but the PGP-encrypted version (.xlsx.gpg)
exists, the script will attempt to decrypt it automatically using gpg.
"""
import os
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl not found, trying to install...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

XLSX_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "DRAFT Working Copy CDG Consensus - PeRef Logical Information Model (Data Dictionary).xlsx"
)

GPG_PATH = XLSX_PATH + ".gpg"

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "input", "data-dictionary"
)


def ensure_decrypted():
    """Decrypt the .xlsx.gpg file if the plaintext .xlsx is missing."""
    if os.path.exists(XLSX_PATH):
        return  # plaintext already available
    if not os.path.exists(GPG_PATH):
        print(f"ERROR: Neither plaintext nor encrypted data dictionary found.\n"
              f"  Expected: {XLSX_PATH}\n"
              f"       or:  {GPG_PATH}", file=sys.stderr)
        sys.exit(1)
    print(f"Plaintext not found — decrypting {os.path.basename(GPG_PATH)} ...")
    result = subprocess.run(
        ["gpg", "--decrypt", "--output", XLSX_PATH, GPG_PATH],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"ERROR: gpg decrypt failed:\n{result.stderr}", file=sys.stderr)
        sys.exit(1)
    print("  Decryption successful.")

def sanitize_sheet_name(name):
    """Make a sheet name safe for use as a filename."""
    return "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in name).strip().replace(' ', '_')

def extract_all_sheets():
    ensure_decrypted()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    
    print(f"Workbook has {len(wb.sheetnames)} sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sanitize_sheet_name(sheet_name)
        csv_path = os.path.join(OUTPUT_DIR, f"{safe_name}.csv")
        
        import csv
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                # Convert None to empty string
                writer.writerow([str(cell) if cell is not None else '' for cell in row])
                row_count += 1
        
        print(f"  Sheet '{sheet_name}' -> {csv_path} ({row_count} rows)")
    
    wb.close()
    print(f"\nDone. CSV files written to {OUTPUT_DIR}")

if __name__ == "__main__":
    extract_all_sheets()
